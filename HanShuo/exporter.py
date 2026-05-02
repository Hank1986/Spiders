import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


COLUMNS = [
    ("caseId",                      "案件ID"),
    ("caseNo",                      "案件编号"),
    ("caseStatus",                  "案件状态码"),
    ("caseStatusText",              "案件状态"),
    ("productName",                 "产品名称"),
    ("userName",                    "借款人姓名"),
    ("idno",                        "身份证号"),
    ("userPhone",                   "手机号"),
    ("handleAmount",                "委案金额"),
    ("caseAlreadyRepaidAmount",     "已还金额"),
    ("toBeRepaidHandleAmount",      "待还委案金额"),
    ("residueAmount",               "剩余金额"),
    ("judicialAdvancesAmount",      "司法垫付金额"),
    ("residueJudicialAdvancesAmount","剩余司法垫付"),
    ("totalRebuildAmount",          "重组总金额"),
    ("loanPactNo",                  "借款合同号"),
    ("orgTitle",                    "机构名称"),
    ("distLogName",                 "分配人"),
    ("distTime",                    "分配时间"),
    ("entrustTime",                 "委案时间"),
    ("allotTime",                   "指派时间"),
    ("deptName",                    "部门"),
    ("cpeName",                     "催收员"),
    ("followStatusText",            "跟进状态"),
    ("entrustContactResultText",    "委案联系结果"),
    ("entrustFollowTimes",          "委案跟进次数"),
    ("entrustLastFollowTime",       "最后跟进时间"),
    ("caseStatusRemark",            "状态备注"),
    ("isSensitive",                 "敏感标记"),
    ("visitStatus",                 "拜访状态"),
]


def export_to_excel(records: list, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "案件列表"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Write header
    for col_idx, (_, cn_name) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=cn_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        for col_idx, (field, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(field))

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)
    print(f"Saved {len(records)} records to {output_path}")
